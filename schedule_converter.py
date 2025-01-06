import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QFileDialog, QLineEdit
from PyQt5.QtCore import Qt
import os
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter


def read_excel_schedule(file_path):
    """从Excel文件读取课程数据"""
    try:
        df = pd.read_excel(file_path)
        return df.to_dict('records')
    except Exception as e:
        print(f"读取Excel文件出错: {e}")
        return None
    
    
def convert_time_to_minutes(time_obj):
    """将时间转换为分钟数，用于排序"""
    try:
        # 处理 datetime.time 对象
        if hasattr(time_obj, 'hour') and hasattr(time_obj, 'minute'):
            return time_obj.hour * 60 + time_obj.minute
            
        # 处理 datetime.datetime 对象
        elif isinstance(time_obj, datetime):
            return time_obj.hour * 60 + time_obj.minute
            
        # 处理字符串格式
        elif isinstance(time_obj, str):
            if ':' in time_obj:
                hours, minutes = map(int, time_obj.split(':'))
                return hours * 60 + minutes
                
        return 0
        
    except Exception as e:
        print(f"时间转换错误: {e}, 时间类型: {type(time_obj)}")
        return 0

def convert_schedule(input_data, student_name):
    """转换课程表格式"""
    # 转换为DataFrame
    odf = pd.DataFrame(input_data)
    
    # 过滤特定学生的课程
    try:
        df = odf[odf['学生姓名'] == student_name]
        print(student_name,df)
        # 将日期转换为datetime对象
        df['日期'] = pd.to_datetime(df['日期'])

        # 获取每周的起始日期
        min_date = df['日期'].min()
        start_of_week = min_date - timedelta(days=min_date.weekday())

        # 创建周课表
        weeks = []
        current_week = start_of_week

        while current_week <= df['日期'].max():
            week_end = current_week + timedelta(days=6)
            week_data = df[(df['日期'] >= current_week) & (df['日期'] <= week_end)]

            if not week_data.empty:
                week_schedule = {
                    'week_start': current_week,
                    'week_end': week_end,
                    'classes': []
                }

                for _, row in week_data.iterrows():
                    class_info = {
                        'weekday': row['日期'].weekday(),
                        'time': row['时间'],
                        'subject': f"{row['科目']}-{row['老师姓名']}"
                    }
                    week_schedule['classes'].append(class_info)

                weeks.append(week_schedule)

            current_week += timedelta(days=7)

        return weeks
    except:
        try:
            
            df = odf[odf['老师姓名'] == student_name]
            print(student_name,df)
            # 将日期转换为datetime对象
            df['日期'] = pd.to_datetime(df['日期'])

            # 获取每周的起始日期
            min_date = df['日期'].min()
            start_of_week = min_date - timedelta(days=min_date.weekday())

            # 创建周课表
            weeks = []
            current_week = start_of_week

            while current_week <= df['日期'].max():
                week_end = current_week + timedelta(days=6)
                week_data = df[(df['日期'] >= current_week) & (df['日期'] <= week_end)]

                if not week_data.empty:
                    week_schedule = {
                        'week_start': current_week,
                        'week_end': week_end,
                        'classes': []
                    }

                    for _, row in week_data.iterrows():
                        class_info = {
                            'weekday': row['日期'].weekday(),
                            'time': row['时间'],
                            'subject': f"{row['科目']}-{row['学生姓名']}"
                        }
                        week_schedule['classes'].append(class_info)

                    weeks.append(week_schedule)

                current_week += timedelta(days=7)

            return weeks
        except:
            print("没找到该名字.")
    
    

def copy_cell_style(source_cell, target_cell):
    target_cell.font = Font(name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            vertAlign=source_cell.font.vertAlign,
                            underline=source_cell.font.underline,
                            strike=source_cell.font.strike,
                            color=source_cell.font.color)

    target_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                   start_color=source_cell.fill.start_color,
                                   end_color=source_cell.fill.end_color)

    target_cell.border = Border(left=source_cell.border.left,
                                right=source_cell.border.right,
                                top=source_cell.border.top,
                                bottom=source_cell.border.bottom)

    target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                      vertical=source_cell.alignment.vertical,
                                      text_rotation=source_cell.alignment.text_rotation,
                                      wrap_text=source_cell.alignment.wrap_text,
                                      shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                      indent=source_cell.alignment.indent)
def format_time(time_obj):
    """格式化时间显示"""
    if isinstance(time_obj, datetime):
        return time_obj.strftime('%H:%M')
    elif hasattr(time_obj, 'strftime'):  # 处理 time 对象
        return time_obj.strftime('%H:%M')
    return str(time_obj)

def export_to_excel(weeks, student_name, output_file):
    """导出为Excel文件"""
    wb = openpyxl.Workbook()
    
    # 设置样式
    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    cell_fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 为每周创建一个工作表
    for week_num, week in enumerate(weeks, 1):
        ws = wb.create_sheet(title=f"Week {week_num}")
        
        # 设置列宽
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
            
        # 添加标题
        week_title = f"{student_name}课程表-WEEK {week_num}"
        ws.merge_cells('A1:H1')
        ws['A1'] = week_title
        ws['A1'].fill = header_fill
        ws['A1'].alignment = align
        
        # 添加日期行
        weekdays = ['时间', '周一', '周二', '周三', '周四', '周五', '周六', '周日']
        dates = [(week['week_start'] + timedelta(days=i)).strftime('(%m%d)') for i in range(7)]
        
        for col, (day, date) in enumerate(zip(weekdays[1:], dates), 2):
            ws.cell(row=2, column=col, value=f"{day}{date}")
            ws.cell(row=2, column=col).fill = header_fill
            ws.cell(row=2, column=col).alignment = align
        
        ws.cell(row=2, column=1, value=weekdays[0])
        ws.cell(row=2, column=1).fill = header_fill
        ws.cell(row=2, column=1).alignment = align
        
        # 创建时间槽字典
        time_slots = {}
        for class_info in week['classes']:
            time = class_info['time']
            if time not in time_slots:
                time_slots[time] = [''] * 7
            time_slots[time][class_info['weekday']] = class_info['subject']
        
        # 填充课程信息
        sorted_times = sorted(time_slots.items(), key=lambda x: convert_time_to_minutes(x[0]))
        for row_num, (time, slots) in enumerate(sorted_times, 3):
            # 使用格式化后的时间
            ws.cell(row=row_num, column=1, value=format_time(time))
            for col_num, content in enumerate(slots, 2):
                cell = ws.cell(row=row_num, column=col_num, value=content)
                cell.fill = cell_fill
                cell.alignment = align
                cell.border = border
        
        # 应用边框到所有使用的单元格
        for row in ws.iter_rows(min_row=1, max_row=len(time_slots)+2, min_col=1, max_col=8):
            for cell in row:
                cell.border = border
    
    # 删除默认的Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
        
    output_ws = wb.create_sheet(title=f"{student_name}课程表")
    current_row = 1

    # 遍历工作簿中的所有工作表
    for sheet in wb.worksheets:
        if sheet.title == f"{student_name}课程表":
            continue  # 跳过合并后的工作表

        # 获取每个工作表的最大行和最大列
        max_row = sheet.max_row
        max_column = sheet.max_column

        # 将当前工作表的内容复制到合并后的工作表中
        for i in range(1, max_row + 1):
            for j in range(1, max_column + 1):
                source_cell = sheet.cell(row=i, column=j)
                target_cell = output_ws.cell(row=current_row, column=j, value=source_cell.value)
                
                # 复制单元格样式
                if source_cell.has_style:
                    copy_cell_style(source_cell, target_cell)
            current_row += 1

        # 在每个工作表内容后添加一个空行
        current_row += 1

    
    # 保存文件
    try:
        wb.save(output_file)
        print(f"课程表已导出到: {output_file}")
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")

class ScheduleConverterGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.input_file = ""
        self.output_file = ""

    def initUI(self):
        self.setWindowTitle('课程表转换器')
        self.setGeometry(300, 300, 500, 200)

        # 创建中心部件和布局
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # 输入文件选择
        input_layout = QHBoxLayout()
        self.input_label = QLabel('输入文件：')
        self.input_path = QLineEdit()
        self.input_button = QPushButton('选择文件')
        self.input_button.clicked.connect(self.select_input_file)
        input_layout.addWidget(self.input_label)
        input_layout.addWidget(self.input_path)
        input_layout.addWidget(self.input_button)

        # 输出文件选择
        output_layout = QHBoxLayout()
        self.output_label = QLabel('输出文件：')
        self.output_path = QLineEdit()
        self.output_button = QPushButton('选择位置')
        self.output_button.clicked.connect(self.select_output_file)
        output_layout.addWidget(self.output_label)
        output_layout.addWidget(self.output_path)
        output_layout.addWidget(self.output_button)

        # 学生姓名输入
        name_layout = QHBoxLayout()
        self.name_label = QLabel('姓名：')
        self.name_input = QLineEdit()
        name_layout.addWidget(self.name_label)
        name_layout.addWidget(self.name_input)

        # 转换按钮
        self.convert_button = QPushButton('开始转换')
        self.convert_button.clicked.connect(self.convert)

        # 状态标签
        self.status_label = QLabel('')
        self.status_label.setAlignment(Qt.AlignCenter)

        # 添加所有组件到主布局
        layout.addLayout(input_layout)
        layout.addLayout(output_layout)
        layout.addLayout(name_layout)
        layout.addWidget(self.convert_button)
        layout.addWidget(self.status_label)

    def select_input_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, '选择输入文件', '', 'Excel Files (*.xlsx *.xls)')
        if file_name:
            self.input_path.setText(file_name)
            self.input_file = file_name

    def select_output_file(self):
        file_name, _ = QFileDialog.getSaveFileName(self, '选择保存位置', '', 'Excel Files (*.xlsx)')
        if file_name:
            self.output_path.setText(file_name)
            self.output_file = file_name

    def convert(self):
        if not self.input_path.text() or not self.output_path.text() or not self.name_input.text():
            self.status_label.setText('请填写所有必要信息！')
            return

        try:
            # 读取输入数据
            input_data = read_excel_schedule(self.input_path.text())
            if input_data is None:
                self.status_label.setText('读取输入文件失败！')
                return

            # 转换课程表
            weeks = convert_schedule(input_data, self.name_input.text())

            # 导出到Excel
            export_to_excel(weeks, self.name_input.text(), self.output_path.text())
            
            self.status_label.setText('转换成功！')
        except Exception as e:
            self.status_label.setText(f'转换失败：{str(e)}')

def main():
    app = QApplication(sys.argv)
    ex = ScheduleConverterGUI()
    ex.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()

